from django.shortcuts import render, redirect, get_object_or_404
from .models import Ubicacion, Armamento, TipoArmamento, Responsable, Movimiento, Mantenimiento, Promocion, Alumno
from .forms import UbicacionForm, ArmamentoForm, TipoArmamentoForm, ResponsableForm, MovimientoForm, MantenimientoForm, FinalizarMantenimientoForm, ReporteArmamentoForm, PromocionForm, AlumnoForm, ReporteAlumnoForm, ImportarMatrizAlumnosForm
from django.contrib import messages
from django.utils import timezone
from datetime import datetime, time

from django.core.paginator import Paginator
from django.db.models import Q, Count, OuterRef, Subquery
from django.db import transaction
from openpyxl import load_workbook
from django.contrib.auth.decorators import login_required, user_passes_test
from .utils import registrar_movimiento

def es_administrador(user):
    return user.groups.filter(name="Administrador").exists()

# Create your views here.
@login_required
@user_passes_test(es_administrador)
def lista_ubicaciones(request):

    buscar = request.GET.get("buscar", "")

    ubicaciones = Ubicacion.objects.all()

    if buscar:

        ubicaciones = ubicaciones.filter(
            Q(codigo__icontains=buscar) |
            Q(nombre__icontains=buscar)
        )

    paginator = Paginator(ubicaciones, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "inventario/ubicaciones/lista.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
        }
    )

@login_required
@user_passes_test(es_administrador)
def crear_ubicacion(request):

    if request.method == "POST":

        form = UbicacionForm(request.POST)

        if form.is_valid():
            form.save()
            messages.success(
                request,
                "La ubicación fue registrada correctamente."
            )
            return redirect("lista_ubicaciones")

    else:

        form = UbicacionForm()

    return render(
        request,
        "inventario/ubicaciones/form.html",
        {
            "form": form,
            "titulo": "Nueva Ubicación"
        }
    )

@login_required
@user_passes_test(es_administrador)
def editar_ubicacion(request, pk):

    ubicacion = get_object_or_404(
        Ubicacion,
        pk=pk
    )

    if request.method == "POST":

        form = UbicacionForm(
            request.POST,
            instance=ubicacion
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "La ubicación fue actualizada correctamente."
            )

            return redirect("lista_ubicaciones")

    else:

        form = UbicacionForm(
            instance=ubicacion
        )

    return render(
        request,
        "inventario/ubicaciones/form.html",
        {
            "form": form,
            "titulo": "Editar Ubicación"
        }
    )

@login_required
@user_passes_test(es_administrador)
def eliminar_ubicacion(request, pk):

    ubicacion = get_object_or_404(
        Ubicacion,
        pk=pk
    )

    if request.method == "POST":

        ubicacion.activo = False
        ubicacion.save()

        messages.success(
            request,
            "La ubicación fue desactivada correctamente."
        )

        return redirect("lista_ubicaciones")

    return render(
        request,
        "inventario/ubicaciones/eliminar.html",
        {
            "ubicacion": ubicacion
        }
    )
#Armamento
@login_required
def lista_armamentos(request):

    buscar = request.GET.get("buscar", "")
    promocion = request.GET.get("promocion")

    ultimo_movimiento_qs = (
        Movimiento.objects
        .filter(
            armamento=OuterRef("pk"),
            tipo__in=["ENTRADA", "SALIDA"]
        )
        .order_by("-fecha")
    )

    armamentos = (
        Armamento.objects
        .filter(activo=True)
        .select_related(
            "tipo",
            "ubicacion",
            "responsable",
            "duenio",
            "duenio__promocion",
        )
        .annotate(
            ultimo_movimiento_qr=Subquery(
                ultimo_movimiento_qs.values("tipo")[:1]
            )
        )
    )

    # ==========================
    # FILTRO POR PROMOCIÓN
    # ==========================

    if promocion:

        armamentos = armamentos.filter(
            duenio__promocion_id=promocion
        )

    # ==========================
    # BÚSQUEDA
    # ==========================

    if buscar:

        armamentos = armamentos.filter(

            Q(codigo__icontains=buscar) |

            Q(numero_serie__icontains=buscar) |

            Q(marca__icontains=buscar) |

            Q(modelo__icontains=buscar)

        )

    # ==========================
    # PROMOCIONES
    # ==========================

    promociones = Promocion.objects.all()

    paginator = Paginator(
        armamentos,
        10
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(
        page_number
    )

    return render(
        request,
        "inventario/armamentos/lista.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
            "promociones": promociones,
            "promocion": promocion,
        }
    )
#Armamento inactivo
@login_required
def armamentos_inactivos(request):

    buscar = request.GET.get("buscar", "")

    armamentos = Armamento.objects.filter(
        activo=False
    ).select_related(
        "tipo",
        "ubicacion",
        "responsable",
        "duenio",
        "duenio__promocion",
    )

    if buscar:

        armamentos = armamentos.filter(
            Q(codigo__icontains=buscar) |
            Q(numero_serie__icontains=buscar) |
            Q(marca__icontains=buscar) |
            Q(modelo__icontains=buscar)
        )

    paginator = Paginator(
        armamentos,
        10
    )

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(
        page_number
    )

    return render(
        request,
        "inventario/armamentos/inactivos.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
        }
    )

#historial armamento
@login_required
def historial_armamento(request, pk):

    armamento = get_object_or_404(
        Armamento,
        pk=pk
    )

    movimientos = (
        Movimiento.objects.select_related(
            "usuario",
            "ubicacion_origen",
            "ubicacion_destino",
            "responsable_anterior",
            "responsable_nuevo",
        )
        .filter(armamento=armamento)
        .order_by("-fecha")
    )

    return render(
        request,
        "inventario/armamentos/historial.html",
        {
            "armamento": armamento,
            "movimientos": movimientos,
        }
    )

@login_required
@user_passes_test(es_administrador)
@transaction.atomic
def crear_armamento(request):

    if request.method == "POST":

        form = ArmamentoForm(request.POST)

        if form.is_valid():

            print("Estado enviado por el formulario:", form.cleaned_data["estado"])

            duenio = form.cleaned_data.get("duenio")

            confirmado = request.POST.get("confirmado")
            print("Dueño seleccionado:", duenio)
            print("Confirmado:", confirmado)

            arma_existente = None

            if duenio:
                arma_existente = Armamento.objects.filter(
                    duenio=duenio,
                    activo=True
                ).first()

            print("Arma encontrada:", arma_existente)

            if duenio and not confirmado:

                arma_existente = Armamento.objects.filter(
                    duenio=duenio,
                    activo=True
                ).first()

                if arma_existente:

                    return render(
                        request,
                        "inventario/armamentos/form.html",
                        {
                            "form": form,
                            "titulo": "Nuevo Armamento",
                            "confirmar_duenio": True,
                            "arma_existente": arma_existente,
                        }
                    )

            armamento = form.save(commit=False)

            print("Estado antes de guardar:", armamento.estado)

            armamento.save()

            print("Estado después de guardar:", armamento.estado)

            registrar_movimiento(

                armamento=armamento,

                tipo="INGRESO",

                usuario=request.user,

                ubicacion_origen=None,
                ubicacion_destino=armamento.ubicacion,

                responsable_anterior=None,
                responsable_nuevo=armamento.responsable,

                estado_anterior=None,
                estado_nuevo=armamento.estado,

                observacion="Registro inicial del armamento."
            )

            messages.success(
                request,
                "Armamento registrado correctamente."
            )

            return redirect("lista_armamentos")

    else:

        form = ArmamentoForm()

    return render(
        request,
        "inventario/armamentos/form.html",
        {
            "form": form,
            "titulo": "Nuevo Armamento"
        }
    )

@login_required
@user_passes_test(es_administrador)
def editar_armamento(request, pk):

    armamento = get_object_or_404(Armamento, pk=pk)

    if request.method == "POST":

        form = ArmamentoForm(
            request.POST,
            instance=armamento
        )

        if form.is_valid():

            armamento_editado = form.save(commit=False)

            # Mantener los valores originales
            armamento_editado.estado = armamento.estado
            armamento_editado.ubicacion = armamento.ubicacion
            armamento_editado.responsable = armamento.responsable

            armamento_editado.save()

            messages.success(
                request,
                "Armamento actualizado correctamente."
            )

            return redirect("lista_armamentos")

    else:

        form = ArmamentoForm(instance=armamento)

    return render(
        request,
        "inventario/armamentos/form.html",
        {
            "form": form,
            "titulo": "Editar Armamento",
        }
    )

@login_required
@user_passes_test(es_administrador)
@transaction.atomic
def eliminar_armamento(request, pk):

    armamento = get_object_or_404(
        Armamento,
        pk=pk
    )

    if request.method == "POST":

        # Guardamos los datos actuales
        estado_anterior = armamento.estado
        ubicacion_anterior = armamento.ubicacion
        responsable_anterior = armamento.responsable
        duenio_anterior = armamento.duenio

        # Baja lógica
        armamento.activo = False

        armamento.save(
            update_fields=["activo"]
        )

        # Registrar la baja en movimientos
        registrar_movimiento(

            armamento=armamento,

            tipo="BAJA",

            usuario=request.user,

            ubicacion_origen=ubicacion_anterior,
            ubicacion_destino=ubicacion_anterior,

            responsable_anterior=responsable_anterior,
            responsable_nuevo=responsable_anterior,

            estado_anterior=estado_anterior,
            estado_nuevo=estado_anterior,

            observacion=(
                "Armamento dado de baja del inventario. "
            )
        )

        messages.success(
            request,
            "El armamento fue dado de baja correctamente."
        )

        return redirect("lista_armamentos")

    return render(
        request,
        "inventario/armamentos/eliminar.html",
        {
            "armamento": armamento
        }
    )

@login_required
@user_passes_test(es_administrador)
def lista_tipos(request):

    buscar = request.GET.get("buscar", "")

    tipos = TipoArmamento.objects.filter(
        Q(nombre__icontains=buscar) |
        Q(descripcion__icontains=buscar)
    )

    paginator = Paginator(tipos, 10)

    page = request.GET.get("page")

    page_obj = paginator.get_page(page)

    return render(request,
        "inventario/tipos/lista.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
        }
    )

@login_required
@user_passes_test(es_administrador)
def crear_tipo(request):

    if request.method == "POST":

        form = TipoArmamentoForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect("lista_tipos")

    else:

        form = TipoArmamentoForm()

    return render(
        request,
        "inventario/tipos/form.html",
        {
            "form": form,
            "titulo": "Nuevo Tipo de Armamento"
        }
    )

@login_required
@user_passes_test(es_administrador)
def editar_tipo(request, pk):

    tipo = get_object_or_404(TipoArmamento, pk=pk)

    if request.method == "POST":

        form = TipoArmamentoForm(request.POST, instance=tipo)

        if form.is_valid():

            form.save()

            return redirect("lista_tipos")

    else:

        form = TipoArmamentoForm(instance=tipo)

    return render(
        request,
        "inventario/tipos/form.html",
        {
            "form": form,
            "titulo": "Editar Tipo de Armamento"
        }
    )

@login_required
@user_passes_test(es_administrador)
def eliminar_tipo(request, pk):

    tipo = get_object_or_404(TipoArmamento, pk=pk)

    if request.method == "POST":

        tipo.delete()

        return redirect("lista_tipos")

    return render(
        request,
        "inventario/tipos/confirmar_eliminar.html",
        {
            "objeto": tipo
        }
    )

@login_required
@user_passes_test(es_administrador)
def lista_responsables(request):

    buscar = request.GET.get("buscar", "")

    responsables = Responsable.objects.all()

    if buscar:
        responsables = responsables.filter(
            Q(cedula__icontains=buscar) |
            Q(nombres__icontains=buscar) |
            Q(apellidos__icontains=buscar)
        )

    paginator = Paginator(responsables, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "inventario/responsables/lista.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
        }
    )

@login_required
@user_passes_test(es_administrador)
def crear_responsable(request):

    if request.method == "POST":

        form = ResponsableForm(request.POST)

        if form.is_valid():

            form.save()

            return redirect("lista_responsables")

    else:

        form = ResponsableForm()

    return render(
        request,
        "inventario/responsables/form.html",
        {
            "form": form,
            "titulo": "Nuevo Responsable"
        }
    )

@login_required
@user_passes_test(es_administrador)
def editar_responsable(request, pk):

    responsable = get_object_or_404(Responsable, pk=pk)

    if request.method == "POST":

        form = ResponsableForm(request.POST, instance=responsable)

        if form.is_valid():

            form.save()

            return redirect("lista_responsables")

    else:

        form = ResponsableForm(instance=responsable)

    return render(
        request,
        "inventario/responsables/form.html",
        {
            "form": form,
            "titulo": "Editar Responsable"
        }
    )

@login_required
@user_passes_test(es_administrador)
def eliminar_responsable(request, pk):

    responsable = get_object_or_404(Responsable, pk=pk)

    if request.method == "POST":

        responsable.delete()

        return redirect("lista_responsables")

    return render(
        request,
        "inventario/responsables/eliminar.html",
        {
            "responsable": responsable
        }
    )

@login_required
def lista_movimientos(request):

    # =====================================================
    # FILTROS
    # =====================================================

    vista = request.GET.get(
        "vista",
        "todos"
    )

    tipo = request.GET.get(
        "tipo",
        ""
    )

    buscar = request.GET.get(
        "buscar",
        ""
    ).strip()

    # =====================================================
    # CONSULTA BASE
    # =====================================================

    movimientos_qs = (
        Movimiento.objects
        .select_related(
            "armamento",
            "usuario",
            "ubicacion_origen",
            "ubicacion_destino",
            "responsable_anterior",
            "responsable_nuevo",
            "duenio_anterior",
            "duenio_nuevo",
        )
        .order_by("-fecha")
    )

    # =====================================================
    # FILTRO PRINCIPAL
    # =====================================================

    if vista == "entradas_salidas":

        movimientos_qs = movimientos_qs.filter(
            tipo__in=[
                "ENTRADA",
                "SALIDA"
            ]
        )

    elif vista == "administrativos":

        movimientos_qs = movimientos_qs.exclude(
            tipo__in=[
                "ENTRADA",
                "SALIDA"
            ]
        )

    # =====================================================
    # FILTRO ENTRADA / SALIDA
    # =====================================================

    if vista == "entradas_salidas":

        if tipo in [
            "ENTRADA",
            "SALIDA"
        ]:

            movimientos_qs = movimientos_qs.filter(
                tipo=tipo
            )

    # =====================================================
    # BUSCAR ARMAMENTO
    # =====================================================

    if buscar:

        movimientos_qs = movimientos_qs.filter(
            Q(
                armamento__codigo__icontains=buscar
            )
            |
            Q(
                armamento__numero_serie__icontains=buscar
            )
        )

    # =====================================================
    # CONTADORES
    # =====================================================

    total_entradas = (
        Movimiento.objects
        .filter(tipo="ENTRADA")
        .count()
    )

    total_salidas = (
        Movimiento.objects
        .filter(tipo="SALIDA")
        .count()
    )

    total_administrativos = (
        Movimiento.objects
        .exclude(
            tipo__in=[
                "ENTRADA",
                "SALIDA"
            ]
        )
        .count()
    )

    # =====================================================
    # CONTEXT
    # =====================================================

    context = {

        "movimientos": movimientos_qs,

        "vista": vista,

        "tipo": tipo,

        "buscar": buscar,

        "total_entradas": total_entradas,

        "total_salidas": total_salidas,

        "total_administrativos":
            total_administrativos,

    }

    return render(
        request,
        "inventario/movimientos/lista.html",
        context
    )

@login_required
@transaction.atomic
def crear_movimiento(request):

    if request.method == "POST":

        form = MovimientoForm(request.POST)

        if form.is_valid():

            movimiento = form.save(commit=False)

            armamento = movimiento.armamento

            movimiento.usuario = request.user

            # Guardar datos anteriores
            movimiento.ubicacion_origen = armamento.ubicacion
            movimiento.responsable_anterior = armamento.responsable
            movimiento.duenio_anterior = armamento.duenio
            movimiento.estado_anterior = armamento.estado

            # ==========================
            # Cambio de ubicación
            # ==========================

            if movimiento.tipo == "CAMBIO_UBICACION":

                armamento.ubicacion = movimiento.ubicacion_destino

            # ==========================
            # Cambio de responsable
            # ==========================

            elif movimiento.tipo == "CAMBIO_RESPONSABLE":

                armamento.responsable = movimiento.responsable_nuevo

            # ==========================
            # Cambio de dueño
            # ==========================

            elif movimiento.tipo == "CAMBIO_DUENIO":

                armamento.duenio = movimiento.duenio_nuevo

            # ==========================
            # No operable
            # ==========================

            elif movimiento.tipo == "NO_OPERABLE":

                armamento.estado = "NO_OPERABLE"

            # Guardar datos nuevos en el movimiento
            movimiento.estado_nuevo = armamento.estado

            armamento.save()

            movimiento.save()

            return redirect("lista_movimientos")

    else:

        form = MovimientoForm()

    return render(
        request,
        "inventario/movimientos/form.html",
        {
            "form": form,
            "titulo": "Registrar Movimiento"
        }
    )

#Generar PDF
from datetime import datetime
from reportlab.lib.units import cm
from .reportes.armamentos_pdf import generar_reporte_armamentos_pdf
from .reportes.responsables_pdf import generar_reporte_responsables_pdf
from .reportes.ubicaciones_pdf import generar_reporte_ubicaciones_pdf
from .reportes.tipos_pdf import generar_reporte_tipos_pdf
from .reportes.movimientos_pdf import generar_reporte_movimientos_pdf
from .reportes.mantenimientos_pdf import generar_reporte_mantenimientos_pdf
from .reportes.promociones_pdf import generar_reporte_promociones_pdf
from .reportes.alumnos_pdf import generar_reporte_alumnos_pdf

# Reporte de armamentos
@login_required
def reporte_armamentos_pdf(request):

    # ==========================================
    # SOLO ARMAMENTOS ACTIVOS
    # ==========================================

    armamentos = Armamento.objects.filter(
        activo=True
    ).select_related(
        "tipo",
        "ubicacion",
        "responsable",
        "duenio",
        "duenio__promocion",
    )

    # ==========================================
    # RESPONSABLE
    # ==========================================

    responsables = request.GET.getlist("responsables")

    if responsables:

        armamentos = armamentos.filter(
            responsable_id__in=responsables
        )

    # ==========================================
    # ARMAMENTOS
    # ==========================================

    armamentos_ids = request.GET.getlist("armamentos")

    if armamentos_ids:

        armamentos = armamentos.filter(
            id__in=armamentos_ids
        )

    # ==========================================
    # ESTADO
    # ==========================================

    estado = request.GET.get("estado")

    if estado:

        armamentos = armamentos.filter(
            estado=estado
        )

    # ==========================================
    # UBICACIÓN
    # ==========================================

    ubicacion = request.GET.get("ubicacion")

    if ubicacion:

        armamentos = armamentos.filter(
            ubicacion_id=ubicacion
        )

    # ==========================================
    # TIPO
    # ==========================================

    tipo = request.GET.get("tipo")

    if tipo:

        armamentos = armamentos.filter(
            tipo_id=tipo
        )

    # ==========================================
    # PROMOCIÓN
    # ==========================================

    promocion = request.GET.get("promocion")

    if promocion:

        armamentos = armamentos.filter(
            duenio__promocion_id=promocion
        )

    # ==========================================
    # GENERAR PDF
    # ==========================================

    return generar_reporte_armamentos_pdf(
        request,
        armamentos
    )

def reporte_responsables_pdf(request):
    return generar_reporte_responsables_pdf(request)

def reporte_ubicaciones_pdf(request):
    return generar_reporte_ubicaciones_pdf(request)

def reporte_tipos_pdf(request):
    return generar_reporte_tipos_pdf(request)

@login_required
def reporte_movimientos_pdf(request):

    if request.method == "POST":

        grado = request.POST.get(
            "firma_grado",
            ""
        ).strip()

        nombres = request.POST.get(
            "firma_nombres",
            ""
        ).strip()

        apellidos = request.POST.get(
            "firma_apellidos",
            ""
        ).strip()

        cargo = request.POST.get(
            "firma_cargo",
            ""
        ).strip()

        if not grado or not nombres or not apellidos or not cargo:

            return render(
                request,
                "inventario/movimientos/firma_pdf_general.html",
                {
                    "error": (
                        "Todos los campos de la firma "
                        "son obligatorios."
                    )
                }
            )

        return generar_reporte_movimientos_pdf(
            request,
            grado,
            nombres,
            apellidos,
            cargo
        )

    return render(
        request,
        "inventario/movimientos/firma_pdf_general.html"
    )


def reporte_mantenimientos_pdf(request):
    return generar_reporte_mantenimientos_pdf(request)

def reporte_promociones_pdf(request):
    return generar_reporte_promociones_pdf(request)

@login_required
def reporte_alumnos_pdf(request):

    alumnos = Alumno.objects.select_related(
        "promocion"
    )

    promocion = request.GET.get("promocion")

    if promocion:

        alumnos = alumnos.filter(
            promocion_id=promocion
        )

    return generar_reporte_alumnos_pdf(
        request,
        alumnos
    )

#Mantenimiento
@login_required
def lista_mantenimientos(request):

    mantenimientos = Mantenimiento.objects.select_related(
        "armamento"
    ).order_by("-fecha_ingreso")

    return render(
        request,
        "inventario/mantenimientos/lista.html",
        {
            "mantenimientos": mantenimientos
        }
    )

@login_required
@transaction.atomic
def crear_mantenimiento(request):

    if request.method == "POST":

        form = MantenimientoForm(request.POST)

        if form.is_valid():

            mantenimiento = form.save(commit=False)

            # Todo mantenimiento nuevo inicia en proceso
            mantenimiento.estado = "EN_PROCESO"

            # Guardar el responsable actual del armamento
            mantenimiento.responsable_armerillo = (
            mantenimiento.armamento.responsable
)

            armamento = mantenimiento.armamento

            # Guardamos los datos anteriores
            estado_anterior = armamento.estado
            ubicacion_anterior = armamento.ubicacion
            responsable_anterior = armamento.responsable

            # Enviar automáticamente al taller
            taller = Ubicacion.objects.get(es_taller=True)

            armamento.estado = "MANTENIMIENTO"
            armamento.ubicacion = taller

            armamento.save()
            mantenimiento.save()

            # Registrar movimiento
            registrar_movimiento(

                armamento=armamento,

                tipo="MANTENIMIENTO",

                usuario=request.user,

                ubicacion_origen=ubicacion_anterior,
                ubicacion_destino=taller,

                responsable_anterior=responsable_anterior,
                responsable_nuevo=responsable_anterior,

                estado_anterior=estado_anterior,
                estado_nuevo="MANTENIMIENTO",

                observacion="Armamento enviado al taller de mantenimiento."
            )

            return redirect("lista_mantenimientos")

    else:

        form = MantenimientoForm()

    return render(
        request,
        "inventario/mantenimientos/form.html",
        {
            "form": form,
            "titulo": "Registrar Mantenimiento"
        }
    )

@login_required
@transaction.atomic
def editar_mantenimiento(request, pk):

    mantenimiento = get_object_or_404(
        Mantenimiento,
        pk=pk
    )

    if mantenimiento.estado == "FINALIZADO":

        messages.warning(
            request,
            "Este mantenimiento ya fue finalizado y no puede modificarse."
        )

        return redirect("lista_mantenimientos")

    if request.method == "POST":

        form = FinalizarMantenimientoForm(
            request.POST,
            instance=mantenimiento
        )

        if form.is_valid():

            mantenimiento = form.save(commit=False)

            armamento = mantenimiento.armamento

            # Datos anteriores
            estado_anterior = armamento.estado
            ubicacion_anterior = armamento.ubicacion
            responsable_anterior = armamento.responsable

            # Si finaliza y no tiene fecha de salida
            if (
                mantenimiento.estado == "FINALIZADO"
                and not mantenimiento.fecha_salida
            ):
                from django.utils import timezone
                mantenimiento.fecha_salida = timezone.now().date()

            # Cambios al armamento
            if mantenimiento.estado == "FINALIZADO":

                armamento.estado = "OPERABLE"

                if mantenimiento.ubicacion_destino:
                    armamento.ubicacion = mantenimiento.ubicacion_destino

                if mantenimiento.responsable_destino:
                    armamento.responsable = mantenimiento.responsable_destino

            elif mantenimiento.estado == "EN_PROCESO":

                taller = Ubicacion.objects.get(es_taller=True)

                armamento.estado = "MANTENIMIENTO"
                armamento.ubicacion = taller

            armamento.save()
            mantenimiento.save()

            # Registrar movimiento
            registrar_movimiento(

                armamento=armamento,

                tipo="MANTENIMIENTO",

                usuario=request.user,

                ubicacion_origen=ubicacion_anterior,
                ubicacion_destino=armamento.ubicacion,

                responsable_anterior=responsable_anterior,
                responsable_nuevo=armamento.responsable,

                estado_anterior=estado_anterior,
                estado_nuevo=armamento.estado,

                observacion=(
                    "Mantenimiento finalizado."
                    if mantenimiento.estado == "FINALIZADO"
                    else "Mantenimiento actualizado."
                )
            )

            return redirect("lista_mantenimientos")

    else:

        form = FinalizarMantenimientoForm(
            instance=mantenimiento
        )

    return render(
        request,
        "inventario/mantenimientos/form.html",
        {
            "form": form,
            "titulo": "Editar Mantenimiento"
        }
    )

@login_required
def eliminar_mantenimiento(request, pk):

    mantenimiento = get_object_or_404(
        Mantenimiento,
        pk=pk
    )

    if mantenimiento.estado == "FINALIZADO":

        messages.warning(
            request,
            "No se puede eliminar un mantenimiento finalizado porque forma parte del historial."
        )

        return redirect("lista_mantenimientos")

    if request.method == "POST":

        mantenimiento.delete()

        return redirect("lista_mantenimientos")

    return render(
        request,
        "inventario/mantenimientos/eliminar.html",
        {
            "mantenimiento": mantenimiento
        }
    )

from django.http import JsonResponse
@login_required
def obtener_responsable_armamento(request, pk):

    armamento = get_object_or_404(
        Armamento.objects.select_related(
            "duenio",
            "duenio__promocion",
            "ubicacion",
            "responsable",
        ),
        pk=pk,
    )

    return JsonResponse({

        "responsable_id": armamento.responsable.id,

        "codigo": armamento.codigo,

        "estado": armamento.get_estado_display(),

        "ubicacion": armamento.ubicacion.nombre,

        "duenio": (
            str(armamento.duenio)
            if armamento.duenio else
            "Sin asignar"
        ),

        "promocion": (
            armamento.duenio.promocion.nombre
            if armamento.duenio else
            "-"
        ),

    })

#Filtro de reportes
@login_required
def reporte_armamentos(request):

    form = ReporteArmamentoForm()

    return render(
        request,
        "inventario/reportes/armamentos_filtro.html",
        {
            "form": form
        }
    )

@login_required
def reporte_alumnos(request):

    form = ReporteAlumnoForm()

    return render(
        request,
        "inventario/reportes/alumnos_filtro.html",
        {
            "form": form
        }
    )

#Promocion
@login_required
def lista_promociones(request):

    buscar = request.GET.get("buscar", "")

    promociones = Promocion.objects.all()

    if buscar:

        promociones = promociones.filter(
            Q(nombre__icontains=buscar) |
            Q(descripcion__icontains=buscar)
        )

    paginator = Paginator(promociones, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "inventario/promociones/lista.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
        }
    )


@login_required
@user_passes_test(es_administrador)
def crear_promocion(request):

    if request.method == "POST":

        form = PromocionForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Promoción registrada correctamente."
            )

            return redirect("lista_promociones")

    else:

        form = PromocionForm()

    return render(
        request,
        "inventario/promociones/form.html",
        {
            "form": form,
            "titulo": "Nueva Promoción"
        }
    )


@login_required
@user_passes_test(es_administrador)
def editar_promocion(request, pk):

    promocion = get_object_or_404(
        Promocion,
        pk=pk
    )

    if request.method == "POST":

        form = PromocionForm(
            request.POST,
            instance=promocion
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Promoción actualizada correctamente."
            )

            return redirect("lista_promociones")

    else:

        form = PromocionForm(instance=promocion)

    return render(
        request,
        "inventario/promociones/form.html",
        {
            "form": form,
            "titulo": "Editar Promoción"
        }
    )


@login_required
@user_passes_test(es_administrador)
def eliminar_promocion(request, pk):

    promocion = get_object_or_404(
        Promocion,
        pk=pk
    )

    if promocion.alumnos.exists():

        messages.error(
            request,
            "No puede eliminar una promoción que tiene alumnos registrados."
        )

        return redirect("lista_promociones")

    if request.method == "POST":

        promocion.delete()

        messages.success(
            request,
            "Promoción eliminada correctamente."
        )

        return redirect("lista_promociones")

    return render(
        request,
        "inventario/promociones/eliminar.html",
        {
            "objeto": promocion,
            "titulo": "Eliminar Promoción"
        }
    )

#Alumno
@login_required
def lista_alumnos(request):

    buscar = request.GET.get("buscar", "")
    promocion_id = request.GET.get("promocion", "")

    alumnos = Alumno.objects.select_related(
        "promocion"
    )

    # Filtro de búsqueda
    if buscar:

        alumnos = alumnos.filter(
            Q(cedula__icontains=buscar) |
            Q(nombres__icontains=buscar) |
            Q(apellidos__icontains=buscar) |
            Q(promocion__nombre__icontains=buscar)
        )

    # Filtro por promoción
    if promocion_id:

        alumnos = alumnos.filter(
            promocion_id=promocion_id
        )

    # Promociones disponibles para el selector
    promociones = Promocion.objects.filter(
        activa=True
    )

    # Paginación
    paginator = Paginator(alumnos, 10)

    page_number = request.GET.get("page")

    page_obj = paginator.get_page(page_number)

    return render(
        request,
        "inventario/alumnos/lista.html",
        {
            "page_obj": page_obj,
            "buscar": buscar,
            "promociones": promociones,
            "promocion_seleccionada": promocion_id,
        }
    )

@login_required
@user_passes_test(es_administrador)
def crear_alumno(request):

    if request.method == "POST":

        form = AlumnoForm(request.POST)

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Alumno registrado correctamente."
            )

            return redirect("lista_alumnos")

    else:

        form = AlumnoForm()

    return render(
        request,
        "inventario/alumnos/form.html",
        {
            "form": form,
            "titulo": "Nuevo Alumno"
        }
    )

@login_required
@user_passes_test(es_administrador)
def editar_alumno(request, pk):

    alumno = get_object_or_404(
        Alumno,
        pk=pk
    )

    if request.method == "POST":

        form = AlumnoForm(
            request.POST,
            instance=alumno
        )

        if form.is_valid():

            form.save()

            messages.success(
                request,
                "Alumno actualizado correctamente."
            )

            return redirect("lista_alumnos")

    else:

        form = AlumnoForm(instance=alumno)

    return render(
        request,
        "inventario/alumnos/form.html",
        {
            "form": form,
            "titulo": "Editar Alumno"
        }
    )

@login_required
@user_passes_test(es_administrador)
def eliminar_alumno(request, pk):

    alumno = get_object_or_404(
        Alumno,
        pk=pk
    )

    if alumno.armamentos.exists():

        messages.error(
            request,
            "No puede eliminar un alumno que tiene armamentos asignados."
        )

        return redirect("lista_alumnos")

    if request.method == "POST":

        alumno.delete()

        messages.success(
            request,
            "Alumno eliminado correctamente."
        )

        return redirect("lista_alumnos")

    return render(
        request,
        "inventario/alumnos/eliminar.html",
        {
            "objeto": alumno,
            "titulo": "Eliminar Alumno"
        }
    )

#Import excel
def separar_apellido_nombre(texto):

    texto = " ".join(str(texto).strip().split())

    partes = texto.split()

    if len(partes) < 2:
        raise ValueError(
            "El campo APELLIDO NOMBRE debe contener al menos "
            "un apellido y un nombre."
        )

    if len(partes) == 2:

        apellidos = partes[0]
        nombres = partes[1]

    elif len(partes) == 3:

        apellidos = " ".join(partes[:2])
        nombres = partes[2]

    else:

        apellidos = " ".join(partes[:2])
        nombres = " ".join(partes[2:])

    return apellidos, nombres
@login_required
@user_passes_test(es_administrador)
def importar_matriz_alumnos(request):

    # ==========================================================
    # CONFIRMAR IMPORTACIÓN
    # ==========================================================

    if request.method == "POST" and request.POST.get(
        "confirmar_importacion"
    ) == "1":

        registros = request.session.get(
            "matriz_alumnos_registros"
        )

        promocion_id = request.session.get(
            "matriz_alumnos_promocion"
        )

        if not registros or not promocion_id:

            messages.error(
                request,
                "No existe una matriz pendiente de importación."
            )

            return redirect("importar_matriz_alumnos")

        try:

            promocion = Promocion.objects.get(
                id=promocion_id,
                activa=True
            )

        except Promocion.DoesNotExist:

            messages.error(
                request,
                "La promoción seleccionada ya no existe o está inactiva."
            )

            return redirect("importar_matriz_alumnos")

        try:

            with transaction.atomic():

                alumnos_creados = 0
                armamentos_creados = 0

                for registro in registros:

                    fila = registro["fila"]

                    # ==================================================
                    # DATOS DEL ALUMNO
                    # ==================================================

                    cedula = registro["cedula"]

                    if not cedula:

                        raise ValueError(
                            f"Fila {fila}: la cédula está vacía."
                        )

                    if len(cedula) != 10 or not cedula.isdigit():

                        raise ValueError(
                            f"Fila {fila}: la cédula "
                            f"'{cedula}' no es válida."
                        )

                    # Verificar cédula duplicada

                    if Alumno.objects.filter(
                        cedula=cedula
                    ).exists():

                        raise ValueError(
                            f"Fila {fila}: la cédula "
                            f"'{cedula}' ya existe."
                        )

                    # ==================================================
                    # SEPARAR APELLIDO Y NOMBRE
                    # ==================================================

                    apellido_nombre = registro[
                        "apellido_nombre"
                    ]

                    try:

                        apellidos, nombres = (
                            separar_apellido_nombre(
                                apellido_nombre
                            )
                        )

                    except ValueError as e:

                        raise ValueError(
                            f"Fila {fila}: {e}"
                        )

                    # ==================================================
                    # CREAR ALUMNO
                    # ==================================================

                    alumno = Alumno.objects.create(

                        promocion=promocion,

                        cedula=cedula,

                        nombres=nombres,

                        apellidos=apellidos,

                        especialidad=registro[
                            "especialidad"
                        ],

                        grado=registro[
                            "grado"
                        ],

                        novedades=registro[
                            "novedades"
                        ],

                        activo=True
                    )

                    alumnos_creados += 1

                    # ==================================================
                    # TIPO DE ARMAMENTO
                    # ==================================================

                    nombre_tipo = registro["tipo"]

                    if not nombre_tipo:

                        raise ValueError(
                            f"Fila {fila}: el tipo de armamento "
                            f"está vacío."
                        )

                    tipo = TipoArmamento.objects.filter(
                        nombre__iexact=nombre_tipo,
                        activo=True
                    ).first()

                    if not tipo:

                        raise ValueError(
                            f"Fila {fila}: no existe el tipo "
                            f"de armamento '{nombre_tipo}'."
                        )

                    # ==================================================
                    # UBICACIÓN
                    # ==================================================

                    nombre_ubicacion = registro[
                        "ubicacion"
                    ]

                    if not nombre_ubicacion:

                        raise ValueError(
                            f"Fila {fila}: la ubicación está vacía."
                        )

                    ubicacion = Ubicacion.objects.filter(
                        nombre__iexact=nombre_ubicacion,
                        activo=True
                    ).first()

                    if not ubicacion:

                        raise ValueError(
                            f"Fila {fila}: no existe la ubicación "
                            f"'{nombre_ubicacion}'."
                        )

                    # ==================================================
                    # RESPONSABLE
                    # ==================================================

                    responsable_texto = registro[
                        "responsable"
                    ]

                    if not responsable_texto:

                        raise ValueError(
                            f"Fila {fila}: el responsable "
                            f"está vacío."
                        )

                    responsable = None

                    responsables = Responsable.objects.filter(
                        activo=True
                    )

                    texto_excel = (
                        responsable_texto
                        .strip()
                        .lower()
                    )

                    for r in responsables:

                        # Formato: APELLIDOS NOMBRES
                        nombre_apellidos = (
                            f"{r.apellidos} {r.nombres}"
                            .strip()
                            .lower()
                        )

                        # Formato: NOMBRES APELLIDOS
                        nombre_nombres = (
                            f"{r.nombres} {r.apellidos}"
                            .strip()
                            .lower()
                        )

                        if texto_excel in [
                            nombre_apellidos,
                            nombre_nombres,
                        ]:

                            responsable = r
                            break

                    if not responsable:

                        raise ValueError(
                            f"Fila {fila}: no se encontró "
                            f"el responsable "
                            f"'{responsable_texto}'."
                        )

                    # ==================================================
                    # DATOS DEL ARMAMENTO
                    # ==================================================

                    codigo = registro["codigo"]

                    numero_serie = registro[
                        "numero_serie"
                    ]

                    if not codigo:

                        raise ValueError(
                            f"Fila {fila}: el código del "
                            f"armamento está vacío."
                        )

                    if not numero_serie:

                        raise ValueError(
                            f"Fila {fila}: el número de serie "
                            f"está vacío."
                        )

                    # Verificar código

                    if Armamento.objects.filter(
                        codigo=codigo
                    ).exists():

                        raise ValueError(
                            f"Fila {fila}: el código "
                            f"'{codigo}' ya existe."
                        )

                    # Verificar número de serie

                    if Armamento.objects.filter(
                        numero_serie=numero_serie
                    ).exists():

                        raise ValueError(
                            f"Fila {fila}: el número de serie "
                            f"'{numero_serie}' ya existe."
                        )

                    # ==================================================
                    # ESTADO
                    # ==================================================

                    estado = registro[
                        "estado"
                    ].upper().strip()

                    estados_validos = [
                        "OPERABLE",
                        "MANTENIMIENTO",
                        "NO_OPERABLE",
                    ]

                    if estado not in estados_validos:

                        raise ValueError(
                            f"Fila {fila}: el estado "
                            f"'{estado}' no es válido. "
                            f"Valores permitidos: "
                            f"{', '.join(estados_validos)}."
                        )

                    # ==================================================
                    # CREAR ARMAMENTO
                    # ==================================================

                    Armamento.objects.create(

                        codigo=codigo,

                        numero_serie=numero_serie,

                        tipo=tipo,

                        marca=registro[
                            "marca"
                        ],

                        modelo=registro[
                            "modelo"
                        ],

                        calibre=registro[
                            "calibre"
                        ],

                        estado=estado,

                        ubicacion=ubicacion,

                        duenio=alumno,

                        responsable=responsable,

                        observaciones=registro[
                            "observaciones"
                        ],

                        activo=True
                    )

                    armamentos_creados += 1

            # ======================================================
            # LIMPIAR SESIÓN
            # ======================================================

            request.session.pop(
                "matriz_alumnos_registros",
                None
            )

            request.session.pop(
                "matriz_alumnos_promocion",
                None
            )

            messages.success(
                request,
                f"Importación completada correctamente. "
                f"Alumnos creados: {alumnos_creados}. "
                f"Armamentos creados: {armamentos_creados}."
            )

            return redirect(
                "lista_alumnos"
            )

        except Exception as e:

            messages.error(
                request,
                f"No se realizó la importación. {e}"
            )

            return redirect(
                "importar_matriz_alumnos"
            )

    # ==========================================================
    # VISTA PREVIA
    # ==========================================================

    if request.method == "POST":

        form = ImportarMatrizAlumnosForm(
            request.POST,
            request.FILES
        )

        if form.is_valid():

            promocion = form.cleaned_data[
                "promocion"
            ]

            archivo = form.cleaned_data[
                "archivo"
            ]

            try:

                workbook = load_workbook(
                    archivo,
                    read_only=True,
                    data_only=True
                )

                hoja = workbook.active

                filas = list(
                    hoja.iter_rows(
                        values_only=True
                    )
                )

                workbook.close()

                if not filas:

                    messages.error(
                        request,
                        "El archivo Excel está vacío."
                    )

                    return render(
                        request,
                        "inventario/alumnos/importar_matriz.html",
                        {
                            "form": form
                        }
                    )

                # ==================================================
                # ENCABEZADOS
                # ==================================================

                encabezados = [
                    str(valor).strip()
                    if valor is not None
                    else ""
                    for valor in filas[0]
                ]

                encabezados_esperados = [
                    "ORD",
                    "ESPECIALIDAD",
                    "GRADO",
                    "APELLIDO NOMBRE",
                    "CEDULA",
                    "N° FUSIL",
                    "NOVEDADES",
                    "NÚMERO DE SERIE",
                    "TIPO",
                    "MARCA",
                    "MODELO",
                    "CALIBRE",
                    "ESTADO",
                    "UBICACIÓN",
                    "RESPONSABLE",
                    "OBSERVACIONES",
                ]

                if encabezados != encabezados_esperados:

                    messages.error(
                        request,
                        "Los encabezados del Excel "
                        "no coinciden con la plantilla."
                    )

                    return render(
                        request,
                        "inventario/alumnos/importar_matriz.html",
                        {
                            "form": form,
                            "encabezados_actuales": encabezados,
                            "encabezados_esperados":
                                encabezados_esperados,
                        }
                    )

                # ==================================================
                # LEER REGISTROS
                # ==================================================

                registros = []

                for numero_fila, fila in enumerate(
                    filas[1:],
                    start=2
                ):

                    if not any(
                        valor is not None
                        and str(valor).strip() != ""
                        for valor in fila
                    ):
                        continue

                    datos = dict(
                        zip(encabezados, fila)
                    )

                    registro = {

                        "fila": numero_fila,

                        "ord": str(
                            datos.get("ORD", "")
                        ).strip(),

                        "especialidad": str(
                            datos.get("ESPECIALIDAD", "")
                        ).strip(),

                        "grado": str(
                            datos.get("GRADO", "")
                        ).strip(),

                        "apellido_nombre": str(
                            datos.get("APELLIDO NOMBRE", "")
                        ).strip(),

                        "cedula": str(
                            datos.get("CEDULA", "")
                        ).strip(),

                        "codigo": str(
                            datos.get("N° FUSIL", "")
                        ).strip(),

                        "novedades": str(
                            datos.get("NOVEDADES", "")
                        ).strip(),

                        "numero_serie": str(
                            datos.get("NÚMERO DE SERIE", "")
                        ).strip(),

                        "tipo": str(
                            datos.get("TIPO", "")
                        ).strip(),

                        "marca": str(
                            datos.get("MARCA", "")
                        ).strip(),

                        "modelo": str(
                            datos.get("MODELO", "")
                        ).strip(),

                        "calibre": str(
                            datos.get("CALIBRE", "")
                        ).strip(),

                        "estado": str(
                            datos.get("ESTADO", "")
                        ).strip().upper(),

                        "ubicacion": str(
                            datos.get("UBICACIÓN", "")
                        ).strip(),

                        "responsable": str(
                            datos.get("RESPONSABLE", "")
                        ).strip(),

                        "observaciones": str(
                            datos.get("OBSERVACIONES", "")
                        ).strip(),
                    }

                    registros.append(
                        registro
                    )

                # ==================================================
                # GUARDAR VISTA PREVIA EN SESIÓN
                # ==================================================

                request.session[
                    "matriz_alumnos_registros"
                ] = registros

                request.session[
                    "matriz_alumnos_promocion"
                ] = promocion.id

                return render(
                    request,
                    "inventario/alumnos/importar_matriz.html",
                    {
                        "form": form,
                        "promocion": promocion,
                        "archivo": archivo,
                        "archivo_recibido": True,
                        "registros": registros,
                        "total_registros": len(
                            registros
                        ),
                    }
                )

            except Exception as e:

                messages.error(
                    request,
                    f"Error al leer el archivo Excel: {e}"
                )

    else:

        form = ImportarMatrizAlumnosForm()

    return render(
        request,
        "inventario/alumnos/importar_matriz.html",
        {
            "form": form
        }
    )
#MAtriz excel
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from django.http import HttpResponse

def descargar_plantilla_matriz(request):

    workbook = Workbook()

    hoja = workbook.active
    hoja.title = "Matriz de Alumnos"

    encabezados = [
        "ORD",
        "ESPECIALIDAD",
        "GRADO",
        "APELLIDO NOMBRE",
        "CEDULA",
        "N° FUSIL",
        "NOVEDADES",
        "NÚMERO DE SERIE",
        "TIPO",
        "MARCA",
        "MODELO",
        "CALIBRE",
        "ESTADO",
        "UBICACIÓN",
        "RESPONSABLE",
        "OBSERVACIONES",
    ]

    hoja.append(encabezados)

    # Formato de encabezados
    for celda in hoja[1]:
        celda.font = Font(bold=True)
        celda.alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

    # Ancho de columnas
    anchos = [
        8, 18, 15, 28, 14, 18, 22, 22,
        18, 15, 15, 15, 18, 22, 30, 30
    ]

    for numero, ancho in enumerate(anchos, start=1):
        hoja.column_dimensions[
            get_column_letter(numero)
        ].width = ancho

    # Congelar encabezados
    hoja.freeze_panes = "A2"

    # Filtro
    hoja.auto_filter.ref = "A1:P1"

    # Segunda hoja con instrucciones
    instrucciones = workbook.create_sheet(
        "Instrucciones"
    )

    instrucciones.append([
        "INSTRUCCIONES PARA LA IMPORTACIÓN"
    ])

    instrucciones.append([
        "No cambie los nombres de los encabezados de la hoja "
        "'Matriz de Alumnos'."
    ])

    instrucciones.append([
        "Complete una fila por cada alumno y su armamento."
    ])

    instrucciones.append([
        "APELLIDO NOMBRE",
        "Escriba primero los apellidos y después los nombres."
    ])

    instrucciones.append([
        "CEDULA",
        "Debe contener 10 dígitos."
    ])

    instrucciones.append([
        "N° FUSIL",
        "Código único del armamento."
    ])

    instrucciones.append([
        "TIPO",
        "Debe coincidir con un tipo de armamento registrado."
    ])

    instrucciones.append([
        "UBICACIÓN",
        "Debe coincidir con una ubicación registrada."
    ])

    instrucciones.append([
        "RESPONSABLE",
        "Debe coincidir con un responsable registrado."
    ])

    instrucciones.append([
        "ESTADO",
        "OPERABLE, MANTENIMIENTO o NO_OPERABLE."
    ])

    instrucciones.column_dimensions["A"].width = 30
    instrucciones.column_dimensions["B"].width = 90

    instrucciones["A1"].font = Font(
        bold=True,
        size=14
    )

    # Preparar respuesta
    respuesta = HttpResponse(
        content_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        )
    )

    respuesta[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="plantilla_matriz_alumnos_armamentos.xlsx"'
    )

    workbook.save(respuesta)

    return respuesta

#Lector QR
import qrcode
from io import BytesIO
from django.http import HttpResponse
from django.http import JsonResponse
@login_required
def lector_qr(request):
    return render(
        request,
        "inventario/movimientos/lector_qr.html"
    )

@login_required
def generar_qr_armamento(request, pk):

    armamento = get_object_or_404(
        Armamento,
        pk=pk
    )

    qr = qrcode.QRCode(
        version=1,
        box_size=10,
        border=4
    )

    qr.add_data(
        armamento.codigo
    )

    qr.make(
        fit=True
    )

    imagen = qr.make_image(
        fill_color="black",
        back_color="white"
    )

    buffer = BytesIO()

    imagen.save(
        buffer,
        format="PNG"
    )

    buffer.seek(0)

    return HttpResponse(
        buffer.getvalue(),
        content_type="image/png"
    )

@login_required
def consultar_armamento_qr(request):

    codigo = request.GET.get("codigo", "").strip()

    if not codigo:
        return JsonResponse({
            "encontrado": False,
            "mensaje": "No se recibió ningún código."
        })

    try:

        armamento = Armamento.objects.select_related(
            "tipo",
            "duenio",
            "duenio__promocion",
            "ubicacion",
            "responsable"
        ).get(
            codigo=codigo,
            activo=True
        )

    except Armamento.DoesNotExist:

        return JsonResponse({
            "encontrado": False,
            "mensaje": (
                f"El armamento con código "
                f"'{codigo}' no existe o se encuentra inactivo."
            )
        })

    # =====================================================
    # BUSCAR ÚLTIMO MOVIMIENTO DE ENTRADA / SALIDA
    # =====================================================

    ultimo_movimiento = (
        Movimiento.objects
        .filter(
            armamento=armamento,
            tipo__in=["ENTRADA", "SALIDA"]
        )
        .order_by("-fecha")
        .first()
    )

    # =====================================================
    # DETERMINAR ACCIÓN PERMITIDA
    # =====================================================

    if ultimo_movimiento is None:

        # El arma nunca ha registrado una salida.
        # Por lo tanto, la primera acción será SALIDA.

        accion_permitida = "SALIDA"

    elif ultimo_movimiento.tipo == "SALIDA":

        # Si salió, ahora solamente puede entrar.

        accion_permitida = "ENTRADA"

    elif ultimo_movimiento.tipo == "ENTRADA":

        # Si entró, ahora solamente puede salir.

        accion_permitida = "SALIDA"

    else:

        accion_permitida = None

    # =====================================================
    # RESPUESTA
    # =====================================================

    return JsonResponse({

        "encontrado": True,

        "accion_permitida": accion_permitida,

        "ultimo_movimiento": (
            ultimo_movimiento.tipo
            if ultimo_movimiento
            else None
        ),

        "armamento": {

            "id": armamento.id,

            "codigo": armamento.codigo,

            "numero_serie": armamento.numero_serie,

            "tipo": str(armamento.tipo),

            "marca": armamento.marca,

            "modelo": armamento.modelo,

            "calibre": armamento.calibre,

            "estado": armamento.get_estado_display(),

            "ubicacion": str(armamento.ubicacion),

            "responsable": str(armamento.responsable),

            "duenio": (
                str(armamento.duenio)
                if armamento.duenio
                else "Sin dueño asignado"
            ),

            "promocion": (
                str(armamento.duenio.promocion)
                if armamento.duenio
                and armamento.duenio.promocion
                else "Sin promoción"
            ),
        }
    })

@login_required
@transaction.atomic
def registrar_movimiento_qr(request):

    if request.method != "POST":
        return JsonResponse({
            "ok": False,
            "mensaje": "Método no permitido."
        }, status=405)

    codigo = request.POST.get("codigo", "").strip()
    tipo = request.POST.get("tipo", "").strip()

    if tipo not in ["ENTRADA", "SALIDA"]:
        return JsonResponse({
            "ok": False,
            "mensaje": "Tipo de movimiento no válido."
        }, status=400)

    try:

        armamento = Armamento.objects.select_related(
            "ubicacion",
            "responsable",
            "duenio",
            "duenio__promocion",
            "tipo"
        ).get(
            codigo=codigo,
            activo=True
        )

    except Armamento.DoesNotExist:

        return JsonResponse({
            "ok": False,
            "mensaje": (
                f"El armamento '{codigo}' no existe "
                "o se encuentra dado de baja."
            )
        }, status=404)

    # ==========================================
    # Último movimiento de entrada/salida
    # ==========================================

    ultimo_movimiento = (
        Movimiento.objects
        .filter(
            armamento=armamento,
            tipo__in=["ENTRADA", "SALIDA"]
        )
        .order_by("-fecha")
        .first()
    )

    # ==========================================
    # Determinar acción permitida
    # ==========================================

    if ultimo_movimiento:

        if ultimo_movimiento.tipo == "SALIDA":
            accion_permitida = "ENTRADA"

        else:
            accion_permitida = "SALIDA"

    else:

        # Nunca ha salido → la primera acción permitida es SALIDA
        accion_permitida = "SALIDA"

    # ==========================================
    # Validar acción solicitada
    # ==========================================

    if tipo != accion_permitida:

        if accion_permitida == "ENTRADA":

            mensaje = (
                "Este armamento se encuentra fuera del armerillo. "
                "La siguiente acción permitida es registrar su ENTRADA."
            )

        else:

            mensaje = (
                "Este armamento se encuentra dentro del armerillo. "
                "La siguiente acción permitida es registrar su SALIDA."
            )

        return JsonResponse({
            "ok": False,
            "mensaje": mensaje,
            "accion_permitida": accion_permitida
        }, status=400)

    # ==========================================
    # Crear movimiento
    # ==========================================

    Movimiento.objects.create(

        armamento=armamento,

        tipo=tipo,

        ubicacion_origen=armamento.ubicacion,

        ubicacion_destino=armamento.ubicacion,

        responsable_anterior=armamento.responsable,

        responsable_nuevo=armamento.responsable,

        duenio_anterior=armamento.duenio,

        duenio_nuevo=armamento.duenio,

        estado_anterior=armamento.estado,

        estado_nuevo=armamento.estado,

        usuario=request.user,

        registrado_qr=True,

        observacion=(
            "Movimiento registrado mediante lector QR."
        )
    )

    # ==========================================
    # Después de registrar:
    # Entrada → siguiente será SALIDA
    # Salida → siguiente será ENTRADA
    # ==========================================

    siguiente_accion = (
        "ENTRADA"
        if tipo == "SALIDA"
        else
        "SALIDA"
    )

    return JsonResponse({

        "ok": True,

        "mensaje": (
            "Entrada registrada correctamente."
            if tipo == "ENTRADA"
            else
            "Salida registrada correctamente."
        ),

        "tipo": tipo,

        "codigo": armamento.codigo,

        "siguiente_accion": siguiente_accion
    })

@login_required
def movimientos_qr(request):

    fecha = request.GET.get(
        "fecha"
    )

    tipo = request.GET.get(
        "tipo",
        ""
    )

    # =====================================================
    # FECHA POR DEFECTO
    # =====================================================

    if not fecha:

        fecha = timezone.localdate().isoformat()

    # =====================================================
    # CONVERTIR FECHA
    # =====================================================

    try:

        fecha_obj = datetime.strptime(
            fecha,
            "%Y-%m-%d"
        ).date()

    except ValueError:

        fecha_obj = timezone.localdate()

        fecha = fecha_obj.isoformat()

    # =====================================================
    # RANGO DEL DÍA
    # =====================================================

    inicio = timezone.make_aware(
        datetime.combine(
            fecha_obj,
            time.min
        )
    )

    fin = timezone.make_aware(
        datetime.combine(
            fecha_obj,
            time.max
        )
    )

    # =====================================================
    # MOVIMIENTOS QR
    # =====================================================

    movimientos = (
        Movimiento.objects
        .filter(
            registrado_qr=True,
            fecha__gte=inicio,
            fecha__lte=fin,
            tipo__in=[
                "ENTRADA",
                "SALIDA"
            ]
        )
        .select_related(
            "armamento",
            "usuario"
        )
        .order_by(
            "-fecha"
        )
    )

    # =====================================================
    # FILTRO POR TIPO
    # =====================================================

    if tipo in [
        "ENTRADA",
        "SALIDA"
    ]:

        movimientos = movimientos.filter(
            tipo=tipo
        )

    # =====================================================
    # TOTALES
    # =====================================================

    total = movimientos.count()

    total_entradas = movimientos.filter(
        tipo="ENTRADA"
    ).count()

    total_salidas = movimientos.filter(
        tipo="SALIDA"
    ).count()

    context = {

        "movimientos": movimientos,

        "fecha": fecha,

        "tipo": tipo,

        "total": total,

        "total_entradas": total_entradas,

        "total_salidas": total_salidas,

    }

    return render(
        request,
        "inventario/movimientos/movimientos_qr.html",
        context
    )
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle
)

@login_required
def reporte_movimientos_qr_pdf(request):

    fecha = request.GET.get(
        "fecha"
    )

    tipo = request.GET.get(
        "tipo",
        ""
    )

    # =====================================================
    # FECHA
    # =====================================================

    if not fecha:

        fecha_obj = timezone.localdate()

        fecha = fecha_obj.isoformat()

    else:

        try:

            fecha_obj = datetime.strptime(
                fecha,
                "%Y-%m-%d"
            ).date()

        except ValueError:

            fecha_obj = timezone.localdate()

            fecha = fecha_obj.isoformat()

    # =====================================================
    # RANGO DEL DÍA
    # =====================================================

    inicio = timezone.make_aware(
        datetime.combine(
            fecha_obj,
            time.min
        )
    )

    fin = timezone.make_aware(
        datetime.combine(
            fecha_obj,
            time.max
        )
    )

    # =====================================================
    # MOVIMIENTOS QR
    # =====================================================

    movimientos = (
        Movimiento.objects
        .filter(
            registrado_qr=True,
            fecha__gte=inicio,
            fecha__lte=fin,
            tipo__in=[
                "ENTRADA",
                "SALIDA"
            ]
        )
        .select_related(
            "armamento",
            "usuario"
        )
        .order_by(
            "fecha"
        )
    )

    # =====================================================
    # FILTRO TIPO
    # =====================================================

    if tipo in [
        "ENTRADA",
        "SALIDA"
    ]:

        movimientos = movimientos.filter(
            tipo=tipo
        )

    # =====================================================
    # CONTADORES
    # =====================================================

    total = movimientos.count()

    total_entradas = movimientos.filter(
        tipo="ENTRADA"
    ).count()

    total_salidas = movimientos.filter(
        tipo="SALIDA"
    ).count()

    # =====================================================
    # RESPUESTA PDF
    # =====================================================

    response = HttpResponse(
        content_type="application/pdf"
    )

    response[
        "Content-Disposition"
    ] = (
        f'inline; filename="movimientos_qr_{fecha}.pdf"'
    )

    # =====================================================
    # DOCUMENTO
    # =====================================================

    documento = SimpleDocTemplate(

        response,

        pagesize=A4,

        rightMargin=1.5 * cm,

        leftMargin=1.5 * cm,

        topMargin=1.5 * cm,

        bottomMargin=1.5 * cm,
    )

    estilos = getSampleStyleSheet()

    elementos = []

    # =====================================================
    # TÍTULO
    # =====================================================

    elementos.append(
        Paragraph(
            "REPORTE DE MOVIMIENTOS QR",
            estilos["Title"]
        )
    )

    elementos.append(
        Spacer(
            1,
            0.4 * cm
        )
    )

    elementos.append(
        Paragraph(
            f"Fecha: {fecha_obj.strftime('%d/%m/%Y')}",
            estilos["Normal"]
        )
    )

    elementos.append(
        Spacer(
            1,
            0.3 * cm
        )
    )

    # =====================================================
    # RESUMEN
    # =====================================================

    resumen = [

        [
            "Total movimientos",
            "Entradas",
            "Salidas"
        ],

        [
            str(total),
            str(total_entradas),
            str(total_salidas)
        ]

    ]

    tabla_resumen = Table(
        resumen,
        colWidths=[
            5 * cm,
            5 * cm,
            5 * cm
        ]
    )

    tabla_resumen.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#212529")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, 0),
                8
            ),

        ])
    )

    elementos.append(
        tabla_resumen
    )

    elementos.append(
        Spacer(
            1,
            0.6 * cm
        )
    )

    # =====================================================
    # TABLA DE MOVIMIENTOS
    # =====================================================

    datos = [

        [
            "Hora",
            "Código",
            "Serie",
            "Movimiento",
            "Usuario"
        ]

    ]

    for movimiento in movimientos:

        datos.append([

            movimiento.fecha.strftime(
                "%H:%M:%S"
            ),

            movimiento.armamento.codigo,

            movimiento.armamento.numero_serie,

            movimiento.get_tipo_display(),

            str(movimiento.usuario),

        ])

    if len(datos) == 1:

        datos.append([

            "-",
            "No existen movimientos",
            "-",
            "-",
            "-"

        ])

    tabla = Table(

        datos,

        colWidths=[

            2.3 * cm,
            3.2 * cm,
            3.2 * cm,
            4 * cm,
            4 * cm

        ],

        repeatRows=1
    )

    tabla.setStyle(
        TableStyle([

            (
                "BACKGROUND",
                (0, 0),
                (-1, 0),
                colors.HexColor("#212529")
            ),

            (
                "TEXTCOLOR",
                (0, 0),
                (-1, 0),
                colors.white
            ),

            (
                "FONTNAME",
                (0, 0),
                (-1, 0),
                "Helvetica-Bold"
            ),

            (
                "GRID",
                (0, 0),
                (-1, -1),
                0.5,
                colors.grey
            ),

            (
                "ALIGN",
                (0, 0),
                (-1, -1),
                "CENTER"
            ),

            (
                "VALIGN",
                (0, 0),
                (-1, -1),
                "MIDDLE"
            ),

            (
                "FONTSIZE",
                (0, 0),
                (-1, -1),
                8
            ),

            (
                "TOPPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

            (
                "BOTTOMPADDING",
                (0, 0),
                (-1, -1),
                5
            ),

        ])
    )

    elementos.append(
        tabla
    )

    # =====================================================
    # GENERAR PDF
    # =====================================================

    documento.build(
        elementos
    )

    return response

# ============================================================
# PDF INDIVIDUAL DE MOVIMIENTO
# ============================================================

from .reportes.reporte_movimiento_individual import (
    generar_reporte_movimiento_individual_pdf
)


@login_required
def reporte_movimiento_individual_pdf(request, movimiento_id):

    try:

        movimiento = (
            Movimiento.objects
            .select_related(
                "armamento",
                "usuario",
                "ubicacion_origen",
                "ubicacion_destino",
                "responsable_anterior",
                "responsable_nuevo",
                "duenio_anterior",
                "duenio_nuevo",
            )
            .get(
                id=movimiento_id
            )
        )

    except Movimiento.DoesNotExist:

        return HttpResponse(
            "El movimiento no existe.",
            status=404
        )

    # ========================================================
    # DETERMINAR SI ES CAMBIO DE RESPONSABLE
    # ========================================================

    es_cambio_responsable = (
        movimiento.tipo == "CAMBIO_RESPONSABLE"
    )

    # ========================================================
    # PROCESAR FORMULARIO
    # ========================================================

    if request.method == "POST":

        # ----------------------------------------------------
        # DATOS FIRMA CENTRAL
        # ----------------------------------------------------

        grado_jefe = request.POST.get(
            "firma_grado",
            ""
        ).strip()

        nombres_jefe = request.POST.get(
            "firma_nombres",
            ""
        ).strip()

        apellidos_jefe = request.POST.get(
            "firma_apellidos",
            ""
        ).strip()

        cargo_jefe = request.POST.get(
            "firma_cargo",
            ""
        ).strip()

        # ----------------------------------------------------
        # VALIDAR FIRMA CENTRAL
        # ----------------------------------------------------

        if (
            not grado_jefe
            or not nombres_jefe
            or not apellidos_jefe
            or not cargo_jefe
        ):

            return render(
                request,
                "inventario/movimientos/firma_pdf_individual.html",
                {
                    "movimiento": movimiento,
                    "es_cambio_responsable": es_cambio_responsable,
                    "error": (
                        "Debe completar todos los datos "
                        "de la firma central."
                    )
                }
            )

        # ====================================================
        # SI ES CAMBIO DE RESPONSABLE
        # ====================================================

        if es_cambio_responsable:

            # ------------------------------------------------
            # FIRMA CUSTODIO ENTREGA
            # ------------------------------------------------

            grado_entrega = request.POST.get(
                "grado_entrega",
                ""
            ).strip()

            nombres_entrega = request.POST.get(
                "nombres_entrega",
                ""
            ).strip()

            apellidos_entrega = request.POST.get(
                "apellidos_entrega",
                ""
            ).strip()

            cargo_entrega = request.POST.get(
                "cargo_entrega",
                ""
            ).strip()

            # ------------------------------------------------
            # FIRMA CUSTODIO RECIBE
            # ------------------------------------------------

            grado_recibe = request.POST.get(
                "grado_recibe",
                ""
            ).strip()

            nombres_recibe = request.POST.get(
                "nombres_recibe",
                ""
            ).strip()

            apellidos_recibe = request.POST.get(
                "apellidos_recibe",
                ""
            ).strip()

            cargo_recibe = request.POST.get(
                "cargo_recibe",
                ""
            ).strip()

            # ------------------------------------------------
            # VALIDAR CUSTODIO ENTREGA
            # ------------------------------------------------

            if (
                not grado_entrega
                or not nombres_entrega
                or not apellidos_entrega
                or not cargo_entrega
            ):

                return render(
                    request,
                    "inventario/movimientos/firma_pdf_individual.html",
                    {
                        "movimiento": movimiento,
                        "es_cambio_responsable": True,
                        "error": (
                            "Debe completar todos los datos "
                            "del custodio que entrega."
                        )
                    }
                )

            # ------------------------------------------------
            # VALIDAR CUSTODIO RECIBE
            # ------------------------------------------------

            if (
                not grado_recibe
                or not nombres_recibe
                or not apellidos_recibe
                or not cargo_recibe
            ):

                return render(
                    request,
                    "inventario/movimientos/firma_pdf_individual.html",
                    {
                        "movimiento": movimiento,
                        "es_cambio_responsable": True,
                        "error": (
                            "Debe completar todos los datos "
                            "del custodio que recibe."
                        )
                    }
                )

            # ------------------------------------------------
            # GENERAR PDF CON 3 FIRMAS
            # ------------------------------------------------

            return generar_reporte_movimiento_individual_pdf(

                request=request,

                movimiento=movimiento,

                # Firma central
                firma_grado=grado_jefe,
                firma_nombres=nombres_jefe,
                firma_apellidos=apellidos_jefe,
                firma_cargo=cargo_jefe,

                # Firma entrega
                entrega_grado=grado_entrega,
                entrega_nombres=nombres_entrega,
                entrega_apellidos=apellidos_entrega,
                entrega_cargo=cargo_entrega,

                # Firma recibe
                recibe_grado=grado_recibe,
                recibe_nombres=nombres_recibe,
                recibe_apellidos=apellidos_recibe,
                recibe_cargo=cargo_recibe,

                tres_firmas=True

            )

        # ====================================================
        # OTROS MOVIMIENTOS: UNA FIRMA CENTRAL
        # ====================================================

        return generar_reporte_movimiento_individual_pdf(

            request=request,

            movimiento=movimiento,

            firma_grado=grado_jefe,
            firma_nombres=nombres_jefe,
            firma_apellidos=apellidos_jefe,
            firma_cargo=cargo_jefe,

            tres_firmas=False

        )

    # ========================================================
    # MOSTRAR FORMULARIO
    # ========================================================

    return render(
        request,
        "inventario/movimientos/firma_pdf_individual.html",
        {
            "movimiento": movimiento,
            "es_cambio_responsable": es_cambio_responsable
        }
    )